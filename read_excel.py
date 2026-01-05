from utils import read_excel_to_nested_dict
import openpyxl
import openpyxl
from openpyxl.utils import get_column_letter

def read_excel_to_nested_dict(file_path):
    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return {}

    nested_dict = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_dict = {}

        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
            col_letter = get_column_letter(col[0].column)
            col_dict = {}

            for cell in col:
                row_num = cell.row

                # Handle merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Find the actual cell that holds the value
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                            value = top_left_cell.value
                            break
                else:
                    value = cell.value

                # Skip empty rows and columns
                if value is not None:
                    col_dict[row_num] = value

            # Skip empty columns
            if col_dict:
                sheet_dict[col_letter] = col_dict

        nested_dict[sheet_name] = sheet_dict

    return nested_dict





excel_sheet = read_excel_to_nested_dict("Sk Pak expenses July 2022 21.24.05.xlsx")['PAK']
coloumns_to_include = ["A", "B", "C"]
rows_to_include = range(5, 69)
processed_list = []
for row in rows_to_include:
    print(row)
    for coloumn in coloumns_to_include:
        print(coloumn)
        
        if coloumn in excel_sheet.keys() and row in excel_sheet[coloumn].keys():
            print(f"      {excel_sheet[coloumn][row]}")
        