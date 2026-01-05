import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


def dict_categories_to_excel_single_sheet(data_dict, file_name):
    """
    Convert a dictionary of categories (each with a list of dictionaries) to a single Excel sheet,
    with each category having its own header row, followed by column headers and then data rows.
    If a category has no items, only the category header and a blank row will be added.

    :param data_dict: Dictionary where each key is a category and each value is a list of dictionaries (or an empty list).
    :param file_name: Name of the output Excel file.
    """
    # Initialize an empty list to hold the data
    combined_data = []

    for category, item_list in data_dict.items():
        if item_list:  # If the item list is not empty
            # Convert the list of dictionaries to a DataFrame
            df: pd.DataFrame = pd.DataFrame(item_list)

            # Create a DataFrame for the category header
            header_df = pd.DataFrame(
                [[category] + [""] * (len(df.columns) - 1)], columns=df.columns
            )

            # Create a DataFrame for the column headers
            column_headers_df = pd.DataFrame([df.columns.tolist()], columns=df.columns)

            # Append the category header row
            combined_data.append(header_df)
            # Append the column headers row
            combined_data.append(column_headers_df)
            # Append the actual data rows
            combined_data.append(df)

        else:  # If the item list is empty
            # Create a simple header with one column for the category name
            header_df = pd.DataFrame(
                [[category] + [""] * (len(df.columns) - 1)], columns=df.columns
            )
            combined_data.append(header_df)

        # Append a blank row as a buffer between categories
        # sub_total_df = pd.DataFrame([["SUBTOTAL"] + ["0"] + [''] * (len(df.columns) - 2)], columns=df.columns)
        # combined_data.append(sub_total_df)
        combined_data.append(
            pd.DataFrame([[""] * len(header_df.columns)], columns=header_df.columns)
        )

    # Concatenate all parts into a single DataFrame
    final_df = pd.concat(combined_data, ignore_index=True)

    # Write the final DataFrame to an Excel file
    final_df.to_excel(file_name, index=False, header=False, engine="openpyxl")


def read_excel_to_nested_dict(file_path):
    # Load the workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return {}

    nested_dict = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_dict = {}

        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
            col_letter = get_column_letter(col[0].column)
            col_dict = {}

            for cell in col:
                row_num = cell.row

                # Handle merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Find the actual cell that holds the value
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = sheet.cell(
                                merged_range.min_row, merged_range.min_col
                            )
                            value = top_left_cell.value
                            break
                else:
                    value = cell.value

                # Skip empty rows and columns
                if value is not None:
                    col_dict[row_num] = value

            # Skip empty columns
            if col_dict:
                sheet_dict[col_letter] = col_dict

        nested_dict[sheet_name] = sheet_dict

    return nested_dict


def list_dicts_to_excel(data_list, file_name):
    """
    Convert a list of dictionaries to an Excel file.

    :param data_list: List of dictionaries where each dictionary represents a row.
    :param file_name: Name of the output Excel file.
    """
    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_list)

    # Write the DataFrame to an Excel file
    df.to_excel(file_name, index=False, engine="openpyxl")


if __name__ == "__main__":
    excel_sheet = read_excel_to_nested_dict("Sk Pak expenses July 2022 21.24.05.xlsx")[
        "PAK"
    ]
    coloumns_to_include = ["A", "B", "C"]
    rows_to_include = range(5, 69)
    processed_list = []
    for row in rows_to_include:
        print(row)
        for coloumn in coloumns_to_include:
            print(coloumn)

            if coloumn in excel_sheet.keys() and row in excel_sheet[coloumn].keys():
                print(f"      {excel_sheet[coloumn][row]}")
